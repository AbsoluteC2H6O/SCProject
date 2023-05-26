import gym
import time
import gym_environments
from gym.envs.registration import register
from agent import MonteCarlo
from agent_deterministic import MonteCarloDet
import matplotlib.pyplot as plt
import numpy as np
import os
from Env.settings import randomCels
from openpyxl import Workbook

# registro Env
register (
    id="RobotMaze-v0",
    entry_point="Env.v0.robot_maze:RobotMazeEnv",
)
register (
    id="FrozenLake-v1",
    entry_point="Env.frozen_lake:FrozenLakeEnv",
)

# Allowing environment to have sounds
if "SDL_AUDIODRIVER" in os.environ:
    del os.environ["SDL_AUDIODRIVER"]


def trainDeterminisctic(env, agent, episodes,total_rewards):
    for _ in range(episodes):
        observation, _ = env.reset()
        terminated, truncated = False, False
        iterator =0
        i=0
        terminatedEp = False
        truncatedEp  = False
        while not (terminated or truncated or iterator>=1000):
            if(iterator ==0):
                action = agent.get_action(observation)
            else:
                action = agent.get_pi_action(observation)
            new_observation, reward, terminated, truncated, _ = env.step(action)
            agent.update(observation, action, reward, terminated)
            observation = new_observation
            total_rewards[i] += iterator
            iterator+=1
            terminatedEp= terminated
            truncatedEp = truncated
            if(iterator>=1000):
                truncatedEp = True
                break
        i+=1
                  
    if(terminatedEp):
        return 0
    if(truncatedEp):
        return 1  
def trainStocastic(env, agent, episodes,total_rewards):
    for _ in range(episodes):
        observation, _ = env.reset()
        terminated, truncated = False, False
        i=0
        iterator = 0
        terminatedEp = False
        truncatedEp  = False
        while not (terminated or truncated or iterator>=1000):
            action = agent.get_action(observation)
            new_observation, reward, terminated, truncated, _ = env.step(action)
            agent.update(observation, action, reward, terminated)
            observation = new_observation
            total_rewards[i] += iterator
            iterator +=1
            terminatedEp= terminated
            truncatedEp = truncated
            if(iterator>=1000):
                truncatedEp = True
                break
        i+=1
        
    if(terminatedEp):
        return 0
    if(truncatedEp):
        return 1
    
def play(env, agent):
    observation, _ = env.reset()
    terminated, truncated = False, False
    while not (terminated or truncated):
        action = agent.get_best_action(observation)
        observation, reward, terminated, truncated, truncated = env.step(action)
        env.render()
       
        time.sleep(1)
        
def totalRewards(n_episodes):
    total_rewards_q = np.zeros(n_episodes)
    total_rewards_dq = np.zeros(n_episodes)
    return total_rewards_q, total_rewards_dq
    
if __name__ == "__main__":
    # env = gym.make("RobotMaze-v0", render_mode="human")
    seedGam, seedEp = 0.05, 0.3
    episodes = 10000
    total_rewards_det = np.zeros(episodes)
    total_rewards_st = np.zeros(episodes)

    # mode = 'deterministic'
    # if(mode == 'deterministic'):
    #     trainDeterminisctic(env, agent_deterministic, episodes=ep)
    #     agent_deterministic.render()
    #     play(env, agent_deterministic)

    # else:
    #     trainStocastic(env, agent, episodes=ep)
    #     agent.render()
    #     play(env, agent)
    rounds =3
    variations = 10
    steps_stocastic = np.zeros((variations,rounds))
    steps_deterministic = np.zeros((variations,rounds))
    print("Numero de episodios:", episodes)
    print('Numero de estados:', randomCels*randomCels)
    # gamma=0.9, epsilon=0.9 valores buenos

    # Generar libro de excel para tabla de datos.
    wb = Workbook()
    ws = wb.active
    
    fila =2
    ws.cell(row=1, column=1, value=str("Determinista")) 
    ws.cell(row=1, column=2, value=str("¿Ganó?")) 
    ws.cell(row=1, column=3, value=str("Estocastico")) 
    ws.cell(row=1, column=4, value=str("¿Ganó?")) 
    ws.cell(row=1, column=5, value=str("Gamma")) 
    ws.cell(row=1, column=6, value=str("Epsilon")) 
    ws.cell(row=1, column=7, value=str("Estados")) 
    ws.cell(row=2, column=7, value=str(randomCels*randomCels))
    for i in range(variations):
        for j in range(rounds):
            print('\nParametros: Epsilon = {}, Gamma = {}\n'.format(seedEp,seedGam ))
            env = gym.make("FrozenLake-v1", render_mode="human")
            agentStocastic = MonteCarlo(
                env.observation_space.n, env.action_space.n, gamma=seedGam, epsilon=seedEp
            )
            agent_deterministic = MonteCarloDet(
                env.observation_space.n, env.action_space.n, gamma=seedGam
            )
            ws.cell(row=fila, column=5, value=str(seedGam)) 
            ws.cell(row=fila, column=6, value=str(seedEp)) 
            # Determinista
            print('Calculo en modo Determinista')
            terminate = trainDeterminisctic(env, agent_deterministic, episodes, total_rewards_det)
            print('\tSteps de episodios determinista',total_rewards_det[0])
            ws.cell(row=fila, column=1, value=str(total_rewards_det[0])) 
            agent_deterministic.reset()
            # Estocastico
            env = gym.make("FrozenLake-v1", render_mode="human")
            print('\nCalculo en modo Estocastico')
            truncate = trainStocastic(env, agentStocastic, episodes, total_rewards_st)
            
            if(truncate == 0):
                ws.cell(row=fila, column=2, value=str("Si")) 
            else:
                ws.cell(row=fila, column=2, value=str("No")) 
                
            if(truncate == 0):
                ws.cell(row=fila, column=4, value=str("Si")) 
            else:
                ws.cell(row=fila, column=4, value=str("No")) 
    
            print('\tSteps de episodios estocastico',total_rewards_st[0])
            ws.cell(row=fila, column=3, value=str(total_rewards_st[0])) 
            env.reset()
            agentStocastic.reset()
            
            steps_stocastic[i][j]=total_rewards_det[0]
            steps_deterministic[i][j]=total_rewards_st[0]
            total_rewards_det, total_rewards_st = totalRewards(n_episodes=episodes)
            seedEp += 0.2
            fila+=1
        seedGam += 0.1
        seedEp = 0.3
    wb.save(r"RobotMaze-Baterry.xlsx")
    print("\nSteps Estocastico\n")
    for k in range(variations):
        print(steps_stocastic[k])
        
    print("\nsteps Determinista\n")
    for l in range(variations):
        print(steps_deterministic[l])
    env.close()